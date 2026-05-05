import marimo

__generated_with = "0.23.3"
app = marimo.App(width="medium")


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ### Question: How can I match Power Plant Generators between 923 and 860 datasets?

    EIA 923: We are looking at the 'Generator Data' sheet within the excel dataset
    """)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ## This marimo notebook is part of inital work to combine the following datasets together:
    - EIA 860: https://www.eia.gov/electricity/data/eia860/
    - EIA 923: https://www.eia.gov/electricity/data/eia923/

    ### What is EIA 860? - WHo and What exists dataset:
    Basically this dataset tells us what the generator is
    - Power Plant Locations and IDS
    - Generator level info
    - Ownership and operators
    - Planned additions, retirements and changes
    - Status (operating, retires and proposed)

    #### Important infomation about 860!
    -




    ### What is EIA 923 - What actually happened dataset:
    This datasets tell us how well it performed
    - Monthly electricity generation
    - Fuel consumption
    - Fuel stocks and deliveries
    - Fuel costs and quality

    #### Important infomation about 923
    - A plant will appear multiple times, with the same Plant ID
    - They may have different Prime movers / Fuel types
    - E.g Plant ID 3 has 5 different rows
    - Data is available monthly do we want this yearly?
    - Between different year the names of the columns slightly change!!!


    #### Columns that are important from EIA (2025)
    - Plant ID -
    - Plant Name
    - Operator Name
    - Operator Id
    - Generator Id
    - Net Generation Year To Date

    ### What is our aim?
    - What is the capacity factor change for each power plant from year 0 - Today

    ### How will this be done?
    - Match the ID from EIA-860 with EIA-
    """)
    return


@app.cell
def _():
    import marimo as mo
    import pandas as pd
    import openpyxl
    import xlrd
    from functools import reduce
    file_path = r"C:\Repository Folder\capacity_factors"
    return file_path, mo, pd


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ### Notebook plan and structure:

    1. Read in all eia 860 and eia 923 spreadsheets from 2008-2024 inclusive
    2. Define subroutines:
        - Clean dataframes
        - Select key columns
        - Merge and calculate capacity factors
    3. Run the file, proivding the following ouputs:
        - eia_wide: Singular dataframe with the capcity factor, net generation and capacity for each genrator per year
        - eia_key: List of dataframe, each providing the capcity factor, net generation and capacity for each generator. And can be called on a year basis e.g eia_key[2010]
    """)
    return


@app.cell
def _(file_path, pd):
    # eia_923_2025 = pd.read_excel(file_path + r"\data\eia-923\2025\EIA923_Schedules_2_3_4_5_M_12_2025_20FEB2026.xlsx", sheet_name = 6, header = 4)
    eia_923_2024 = pd.read_excel(file_path + r"\data\eia-923\2024\EIA923_Schedules_2_3_4_5_M_12_2024_Final.xlsx", sheet_name = 8, header = 5)
    eia_923_2023 = pd.read_excel(file_path + r"\data\eia-923\2023\EIA923_Schedules_2_3_4_5_M_12_2023_Final_Revision.xlsx", sheet_name = 8, header = 5)
    eia_923_2022 = pd.read_excel(file_path + r"\data\eia-923\2022\EIA923_Schedules_2_3_4_5_M_12_2022_Final_Revision.xlsx", sheet_name = 7, header = 5) 
    eia_923_2021 = pd.read_excel(file_path + r"\data\eia-923\2021\EIA923_Schedules_2_3_4_5_M_12_2021_Final_Revision.xlsx", sheet_name = 8, header = 5)
    eia_923_2020 = pd.read_excel(file_path + r"\data\eia-923\2020\EIA923_Schedules_2_3_4_5_M_12_2020_Final_Revision.xlsx", sheet_name = 8, header = 5)
    eia_923_2019 = pd.read_excel(file_path + r"\data\eia-923\2019\EIA923_Schedules_2_3_4_5_M_12_2019_Final_Revision.xlsx", sheet_name = 8, header = 5)
    eia_923_2018 = pd.read_excel(file_path + r"\data\eia-923\2018\EIA923_Schedules_2_3_4_5_M_12_2018_Final_Revision.xlsx", sheet_name = 8, header = 5)
    eia_923_2017 = pd.read_excel(file_path + r"\data\eia-923\2017\EIA923_Schedules_2_3_4_5_M_12_2017_Final_Revision.xlsx", sheet_name = 8, header = 5)
    eia_923_2016 = pd.read_excel(file_path + r"\data\eia-923\2016\EIA923_Schedules_2_3_4_5_M_12_2016_Final_Revision.xlsx", sheet_name = 7, header = 5)
    eia_923_2015 = pd.read_excel(file_path + r"\data\eia-923\2015\EIA923_Schedules_2_3_4_5_M_12_2015_Final_Revision.xlsx", sheet_name = 7, header = 5)
    eia_923_2014 = pd.read_excel(file_path + r"\data\eia-923\2014\EIA923_Schedules_2_3_4_5_M_12_2014_Final_Revision.xlsx", sheet_name = 7, header = 5)
    eia_923_2013 = pd.read_excel(file_path + r"\data\eia-923\2013\EIA923_Schedules_2_3_4_5_2013_Final_Revision.xlsx", sheet_name = 6, header = 5)
    eia_923_2012 = pd.read_excel(file_path + r"\data\eia-923\2012\EIA923_Schedules_2_3_4_5_M_12_2012_Final_Revision.xlsx", sheet_name = 6, header = 5)
    eia_923_2011 = pd.read_excel(file_path + r"\data\eia-923\2011\EIA923_Schedules_2_3_4_5_2011_Final_Revision.xlsx", sheet_name = 6, header = 5)
    eia_923_2010 = pd.read_excel(file_path + r"\data\eia-923\2010\EIA923 SCHEDULES 2_3_4_5 Final 2010.xls", sheet_name = 6, header = 7)
    eia_923_2009 = pd.read_excel(file_path + r"\data\eia-923\2009\EIA923 SCHEDULES 2_3_4_5 M Final 2009 REVISED 05252011.XLS", sheet_name = 6, header = 7)
    eia_923_2008 = pd.read_excel(file_path + r"\data\eia-923\2008\eia923December2008.xls", sheet_name = 6, header = 7)

    eia_860_2024 = pd.read_excel(file_path + r"\data\eia-860\2024\3_1_Generator_Y2024.xlsx", sheet_name = 0, header = 1)
    eia_860_2023 = pd.read_excel(file_path + r"\data\eia-860\2023\3_1_Generator_Y2023.xlsx", sheet_name = 0, header = 1)
    eia_860_2022 = pd.read_excel(file_path + r"\data\eia-860\2022\3_1_Generator_Y2022.xlsx", sheet_name = 0, header = 1)
    eia_860_2021 = pd.read_excel(file_path + r"\data\eia-860\2021\3_1_Generator_Y2021.xlsx", sheet_name = 0, header = 1)
    eia_860_2020 = pd.read_excel(file_path + r"\data\eia-860\2020\3_1_Generator_Y2020.xlsx", sheet_name = 0, header = 1)
    eia_860_2019 = pd.read_excel(file_path + r"\data\eia-860\2019\3_1_Generator_Y2019.xlsx", sheet_name = 0, header = 1)
    eia_860_2018 = pd.read_excel(file_path + r"\data\eia-860\2018\3_1_Generator_Y2018.xlsx", sheet_name = 0, header = 1)
    eia_860_2017 = pd.read_excel(file_path + r"\data\eia-860\2017\3_1_Generator_Y2017.xlsx", sheet_name = 0, header = 1)
    eia_860_2016 = pd.read_excel(file_path + r"\data\eia-860\2016\3_1_Generator_Y2016.xlsx", sheet_name = 0, header = 1)
    eia_860_2015 = pd.read_excel(file_path + r"\data\eia-860\2015\3_1_Generator_Y2015.xlsx", sheet_name = 0, header = 1)
    eia_860_2014 = pd.read_excel(file_path + r"\data\eia-860\2014\3_1_Generator_Y2014.xlsx", sheet_name = 0, header = 1)
    eia_860_2013 = pd.read_excel(file_path + r"\data\eia-860\2013\3_1_Generator_Y2013.xlsx", sheet_name = 0, header = 1)
    eia_860_2012 = pd.read_excel(file_path + r"\data\eia-860\2012\GeneratorY2012.xlsx", sheet_name = 0, header = 1)
    eia_860_2011 = pd.read_excel(file_path + r"\data\eia-860\2011\GeneratorY2011.xlsx", sheet_name = 0, header = 1)
    eia_860_2010 = pd.read_excel(file_path + r"\data\eia-860\2010\GeneratorsY2010.xls", sheet_name = 0, header = 0) ## heading tables change
    eia_860_2009 = pd.read_excel(file_path + r"\data\eia-860\2009\GeneratorY09.xls", sheet_name = 0, header = 0)
    eia_860_2008 = pd.read_excel(file_path + r"\data\eia-860\2008\GenY08.xls", sheet_name = 0, header = 0)
    # eia_860_2007 = pd.read_excel(file_path + r"\data\eia-860\2007\GenY07.xls", sheet_name = 0, header = 0)
    # eia_860_2006 = pd.read_excel(file_path + r"\data\eia-860\2006\GenY06.xls", sheet_name = 0, header = 0)
    # eia_860_2005 = pd.read_excel(file_path + r"\data\eia-860\2005\GenY05.xls", sheet_name = 0, header = 0)
    # eia_860_2004 = pd.read_excel(file_path + r"\data\eia-860\2004\GenY04.xls", sheet_name = 0, header = 0)

    return (
        eia_860_2008,
        eia_860_2009,
        eia_860_2010,
        eia_860_2011,
        eia_860_2012,
        eia_860_2013,
        eia_860_2014,
        eia_860_2015,
        eia_860_2016,
        eia_860_2017,
        eia_860_2018,
        eia_860_2019,
        eia_860_2020,
        eia_860_2021,
        eia_860_2022,
        eia_860_2023,
        eia_860_2024,
        eia_923_2008,
        eia_923_2009,
        eia_923_2010,
        eia_923_2011,
        eia_923_2012,
        eia_923_2013,
        eia_923_2014,
        eia_923_2015,
        eia_923_2016,
        eia_923_2017,
        eia_923_2018,
        eia_923_2019,
        eia_923_2020,
        eia_923_2021,
        eia_923_2022,
        eia_923_2023,
        eia_923_2024,
    )


@app.cell
def _(file_path, pd):
    eia_860_retire_2024 = pd.read_excel(file_path + r"\data\eia-860\2024\3_1_Generator_Y2024.xlsx", sheet_name = 2, header = 1)
    eia_860_retire_2023 = pd.read_excel(file_path + r"\data\eia-860\2023\3_1_Generator_Y2023.xlsx", sheet_name = 2, header = 1)
    eia_860_retire_2022 = pd.read_excel(file_path + r"\data\eia-860\2022\3_1_Generator_Y2022.xlsx", sheet_name = 2, header = 1)
    eia_860_retire_2021 = pd.read_excel(file_path + r"\data\eia-860\2021\3_1_Generator_Y2021.xlsx", sheet_name = 2, header = 1)
    eia_860_retire_2020 = pd.read_excel(file_path + r"\data\eia-860\2020\3_1_Generator_Y2020.xlsx", sheet_name = 2, header = 1)
    eia_860_retire_2019 = pd.read_excel(file_path + r"\data\eia-860\2019\3_1_Generator_Y2019.xlsx", sheet_name = 2, header = 1)
    eia_860_retire_2018 = pd.read_excel(file_path + r"\data\eia-860\2018\3_1_Generator_Y2018.xlsx", sheet_name = 2, header = 1)
    eia_860_retire_2017 = pd.read_excel(file_path + r"\data\eia-860\2017\3_1_Generator_Y2017.xlsx", sheet_name = 2, header = 1)
    eia_860_retire_2016 = pd.read_excel(file_path + r"\data\eia-860\2016\3_1_Generator_Y2016.xlsx", sheet_name = 2, header = 1)
    eia_860_retire_2015 = pd.read_excel(file_path + r"\data\eia-860\2015\3_1_Generator_Y2015.xlsx", sheet_name = 2, header = 1)
    eia_860_retire_2014 = pd.read_excel(file_path + r"\data\eia-860\2014\3_1_Generator_Y2014.xlsx", sheet_name = 2, header = 1)
    eia_860_retire_2013 = pd.read_excel(file_path + r"\data\eia-860\2013\3_1_Generator_Y2013.xlsx", sheet_name = 2, header = 1)
    eia_860_retire_2012 = pd.read_excel(file_path + r"\data\eia-860\2012\GeneratorY2012.xlsx", sheet_name = 2, header = 1)
    eia_860_retire_2011 = pd.read_excel(file_path + r"\data\eia-860\2011\GeneratorY2011.xlsx", sheet_name = 2, header = 1)
    eia_860_retire_2010 = pd.read_excel(file_path + r"\data\eia-860\2010\GeneratorsY2010.xls", sheet_name = 2, header = 0) ## heading tables change
    eia_860_retire_2009 = pd.read_excel(file_path + r"\data\eia-860\2009\GeneratorY09.xls", sheet_name = 2, header = 0)
    return (
        eia_860_retire_2009,
        eia_860_retire_2010,
        eia_860_retire_2011,
        eia_860_retire_2012,
        eia_860_retire_2013,
        eia_860_retire_2014,
        eia_860_retire_2015,
        eia_860_retire_2016,
        eia_860_retire_2017,
        eia_860_retire_2018,
        eia_860_retire_2019,
        eia_860_retire_2020,
        eia_860_retire_2021,
        eia_860_retire_2022,
        eia_860_retire_2023,
        eia_860_retire_2024,
    )


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    #### This cell provides the subroutines to clean, select key columns and calculate capacity factors
    """)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    # EIA 860 Pipeline:
    1. Load individual 860 excels as pandas DF
    2. Select the key columns for each year for each generator
    3. From the recently available year select: Retirement year (if possible), nominal capacity, year installed. If a column is blank use the previous year
    """)
    return


@app.function
def clean_columns(df):
    df.columns = (
        df.columns
        .str.lower()
        .str.replace("\n", " ", regex=True)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )
    return df


@app.cell
def _(
    eia_860_2008,
    eia_860_2009,
    eia_860_2010,
    eia_860_2011,
    eia_860_2012,
    eia_860_2013,
    eia_860_2014,
    eia_860_2015,
    eia_860_2016,
    eia_860_2017,
    eia_860_2018,
    eia_860_2019,
    eia_860_2020,
    eia_860_2021,
    eia_860_2022,
    eia_860_2023,
    eia_860_2024,
    pd,
):
    important_columns_860 = (
        "plant code",
        "generator id",
        "nameplate capacity (mw)",
        "technology",
        "operating year" # Year that the generator installed/satrted
    )

    eia_860_years = {
        2024: eia_860_2024,
        2023: eia_860_2023,
        2022: eia_860_2022,
        2021: eia_860_2021,
        2020: eia_860_2020,
        2019: eia_860_2019,
        2018: eia_860_2018,
        2017: eia_860_2017,
        2016: eia_860_2016,
        2015: eia_860_2015,
        2014: eia_860_2014,
        2013: eia_860_2013,
        2012: eia_860_2012,
        2011: eia_860_2011,
        2010: eia_860_2010,
        2009: eia_860_2009,
        2008: eia_860_2008
    }

    fuel_type = {
        "bio energy": ["Wood/Wood Waste Biomass", "Municipal Solid Waste","Landfill Gas","Other Waste Biomass"],
        "coal": ['Conventional Steam Coal','Coal Integrated Gasification Combined Cycle'],
        "gas": ['Natural Gas Steam Turbine','Natural Gas Fired Combined Cycle','Natural Gas Fired Combustion Turbine','Natural Gas Internal Combustion Engine','Natural Gas with Compressed Air Storage','Other Natural Gas'],
        "geothermal": ["Geothermal"],
        "hydro": ["Conventional Hydroelectric"],
        "nuclear": ["Nuclear"],
        "oil": ["Petroleum Liquids"],
        "solar": ["Solar Photovoltaic"],
        "wind onshore": ["Onshore Wind"],
        "wind offshore": ["Offshore Wind"]
    }

    def get_fuel_type(technology):
        for fuel, tech_list in fuel_type.items():
            if technology in tech_list:
                return fuel



    def process_860(df, year):
        df = clean_columns(df)

        if "technology" not in df.columns:
            df["technology"] = "unknown"

        if year <= 2011:
            if year <= 2008:
                df = df.rename(columns={
                    "plntcode": "plant code",
                    "gencode": "generator id",
                    "nameplate": "nameplate capacity (mw)",
                    "operating_year": "operating year"
                })
            else:
                df = df.rename(columns={
                    "generator_id": "generator id",
                    "plant_code": "plant code",
                    "nameplate": "nameplate capacity (mw)",
                    "operating_year": "operating year"
                })
        df = df.loc[:, df.columns.isin(important_columns_860)].copy()

        df = df.rename(columns={
            "plant code": "plant id"#,
            # "nameplate capacity (mw)": ("nameplate capacity (mw) " + str(year))
        })
        df['generator id'] = df['generator id'].astype(str)
        df['reported year'] =  int(year) # Year on the report
        df["operating year"] = pd.to_numeric(df["operating year"], errors="coerce")
        df = df[df["operating year"].notna()]
        df['fuel type'] = df["technology"].apply(get_fuel_type)
        return df

    def fill_missing_years(all_df):
        generators = all_df[["plant id", "generator id"]].drop_duplicates()
        years_df = pd.DataFrame({"reported year": sorted(eia_860_years.keys())})
        idx_df = generators.merge(years_df, how="cross")

        all_df = idx_df.merge(
            all_df,
            on=["plant id", "generator id", "reported year"],
            how="left"
        )

        all_df = all_df.drop_duplicates(
            subset=["plant id", "generator id", "reported year"]
        )

        all_df = all_df.sort_values(
            ["plant id", "generator id", "reported year"]
        )

        group_cols = ["plant id", "generator id"]

        all_df["operating year"] = (
            all_df.groupby(group_cols)["operating year"].transform("min")
        )

        all_df["nameplate capacity (mw)"] = (
            all_df.groupby(group_cols)["nameplate capacity (mw)"]
            .transform(lambda x: x.ffill().bfill()).fillna(0)
        )

        all_df["fuel type"]= (
            all_df.groupby(group_cols)["fuel type"]
            .transform(lambda x: x.ffill().bfill().fillna("unknown"))
        )

        all_df = all_df[ 
            all_df["reported year"] > all_df["operating year"] ## This logic needs to be checked
            ### Potentially can work from the first full year
            ### Could calculate on the month by month
        ]

        all_df = all_df.drop(columns = ["technology"])
        return all_df


    return eia_860_years, fill_missing_years, process_860


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    # EIA 860 Retirement pipeline
    """)
    return


@app.cell
def _(
    eia_860_retire_2009,
    eia_860_retire_2010,
    eia_860_retire_2011,
    eia_860_retire_2012,
    eia_860_retire_2013,
    eia_860_retire_2014,
    eia_860_retire_2015,
    eia_860_retire_2016,
    eia_860_retire_2017,
    eia_860_retire_2018,
    eia_860_retire_2019,
    eia_860_retire_2020,
    eia_860_retire_2021,
    eia_860_retire_2022,
    eia_860_retire_2023,
    eia_860_retire_2024,
    pd,
):
    retirement_column = (
        'plant code',
        'generator id',
        'retirement year'
    )

    retirement_years = {
        2024: eia_860_retire_2024,
        2023: eia_860_retire_2023,
        2022: eia_860_retire_2022,
        2021: eia_860_retire_2021,
        2020: eia_860_retire_2020,
        2019: eia_860_retire_2019,
        2018: eia_860_retire_2018,
        2017: eia_860_retire_2017,
        2016: eia_860_retire_2016,
        2015: eia_860_retire_2015,
        2014: eia_860_retire_2014,
        2013: eia_860_retire_2013,
        2012: eia_860_retire_2012,
        2011: eia_860_retire_2011,
        2010: eia_860_retire_2010,
        2009: eia_860_retire_2009
    }

    def process_retirement(df, year):
        df = clean_columns(df)
        if year <= 2011:
            df = df.rename(columns={
                "generator_id": "generator id",
                "plant_code": "plant code",
                "retirement_year": "retirement year"
            })
        df = df.loc[:, df.columns.isin(retirement_column)].copy()

        df = df.rename(columns={
            "plant code": "plant id"#,
            # "nameplate capacity (mw)": ("nameplate capacity (mw) " + str(year))
        })
        df['generator id'] = df['generator id'].astype(str)
        df['reported year'] =  int(year)
        df = df[
            df['retirement year'].notna() &
            (df["retirement year"].astype(str).str.strip() != "") &
            (pd.to_numeric(df['retirement year'], errors = "coerce") >= 1800)
        ]
        return df

    def combine_retirement(retire, df_860):
        df_860 = df_860.merge(retire, on = ["plant id", "generator id", "reported year"], how = "left")
        return df_860


    return combine_retirement, process_retirement, retirement_years


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    # EIA 923 pipeline
    """)
    return


@app.cell
def _(
    eia_923_2008,
    eia_923_2009,
    eia_923_2010,
    eia_923_2011,
    eia_923_2012,
    eia_923_2013,
    eia_923_2014,
    eia_923_2015,
    eia_923_2016,
    eia_923_2017,
    eia_923_2018,
    eia_923_2019,
    eia_923_2020,
    eia_923_2021,
    eia_923_2022,
    eia_923_2023,
    eia_923_2024,
):
    eia_923_years = {
        #2025: eia_923_2025,
        2024: eia_923_2024,
        2023: eia_923_2023,
        2022: eia_923_2022,
        2021: eia_923_2021,
        2020: eia_923_2020,
        2019: eia_923_2019,
        2018: eia_923_2018,
        2017: eia_923_2017,
        2016: eia_923_2016,
        2015: eia_923_2015,
        2014: eia_923_2014,
        2013: eia_923_2013,
        2012: eia_923_2012,
        2011: eia_923_2011,
        2010: eia_923_2010,
        2009: eia_923_2009,
        2008: eia_923_2008,
    }

    important_columns_923 = (
        "plant id",
        "generator id",
        "net generation year to date"
    )

    def process_923(df, year):
        df = clean_columns(df)
        df = df.loc[:, df.columns.isin(important_columns_923)].copy()
        df['generator id'] = df['generator id'].astype(str)
        df['plant id'] = df['plant id'].astype(int)
        df['reported year'] =  int(year)
        df = df.drop_duplicates(subset=["plant id", "generator id"])
        return df

    return eia_923_years, process_923


@app.cell
def _():
    def merge_and_calculate_cf(df_860, df_923):
        df = df_860.merge(df_923, on = ("plant id", "generator id", "reported year"), how = "left")
        df = df.fillna(0)
        ### Flag missing data
        df['capacity factor'] = df['net generation year to date'] / (df["nameplate capacity (mw)"] * 8760)
        return df
    def calculate_generation_year(df):
        df['generation year'] = df['reported year'] - df['operating year']
        return df

    return calculate_generation_year, merge_and_calculate_cf


@app.cell
def _(
    calculate_generation_year,
    combine_retirement,
    eia_860_years,
    eia_923_years,
    fill_missing_years,
    merge_and_calculate_cf,
    pd,
    process_860,
    process_923,
    process_retirement,
    retirement_years,
):
    eia_860_key = {
        year: process_860(df, year)
        for year, df in eia_860_years.items()
    }

    retirement_key = {
        year: process_retirement(df, year)
        for year, df in retirement_years.items()
    }


    #_test = process_923(eia_923_2020, 2020)
    eia_923_key = {
        year: process_923(df, year)
        for year, df in eia_923_years.items()
    }

    all_860 = pd.concat(eia_860_key.values(), ignore_index=True)
    big_860 = fill_missing_years(all_860)
    all_retirement = pd.concat(retirement_key.values(), ignore_index=True)
    completed_860 = combine_retirement(all_retirement, big_860)
    all_923 = pd.concat(eia_923_key.values(), ignore_index=True)
    eia_complete = merge_and_calculate_cf(completed_860, all_923)
    eia_complete = calculate_generation_year(eia_complete)
    return (eia_complete,)


@app.cell
def _(eia_complete):
    eia_complete
    return


@app.cell
def _(eia_complete):
    eia_complete[eia_complete['reported year'] > eia_complete['retirement year']]
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ### Missing data flag is needed
    """)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ### This section is to create a large dataset to calculate when a generator retired

    - There is a seperate sheet wihtin 860 with retirement documentation which will be able to be calcualted from
    """)
    return


if __name__ == "__main__":
    app.run()
